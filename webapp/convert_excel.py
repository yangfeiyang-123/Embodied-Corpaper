#!/usr/bin/env python3
"""Convert 具身智能文献阅读记录表.xlsx to JSON for webapp import."""
import json, sys, os
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl -q")
    import openpyxl

# Map category sheet names to category values
CATEGORY_MAP = {
    '具身感知与状态估计': '具身感知与状态估计',
    '具身表征与多模态 Grounding': '具身表征与多模态 Grounding',
    'VLA 与机器人基础模型': 'VLA 与机器人基础模型',
    '世界模型与物理后果预测': '世界模型与物理后果预测',
    '策略学习与技能生成': '策略学习与技能生成',
    '接触丰富操作与灵巧控制': '接触丰富操作与灵巧控制',
    '仿真、数据引擎与 Sim-to-Real': '仿真、数据引擎与 Sim-to-Real',
    '评测、部署、安全与泛化': '评测、部署、安全与泛化',
}

FIELD_MAP = [
    ('week', '周次'),
    ('recorder', '记录人'),
    ('readDate', '阅读日期'),
    ('shared', '是否分享'),
    # ('score_overall', '综合评分'),  # computed, skip import
    ('relevance', '相关性(1-5)'),
    ('novelty', '新颖性(1-5)'),
    ('evidence', '证据强度(1-5)'),
    ('inspiration', '启发性(1-5)'),
    ('reproducibility', '可复现性(1-5)'),
    ('title', '论文标题'),
    ('source', '年份/来源'),
    ('link', '链接/项目页'),
    ('direction', '方向'),
    ('oneSentence', '一句话定位'),
    ('authors', '作者与机构'),
    ('task', '核心任务'),
    ('motivation', '问题与动机'),
    ('dataset', '数据集'),
    ('platform', '平台'),
    ('signalAnalysis', '信号分析处理'),
    ('methodOverview', '方法框架：总体架构'),
    ('methodDetails', '方法框架：具体细节和意义目的'),
    ('trainFlow', '实验设计：训练流程'),
    ('hardware', '实验设计：硬件平台'),
    ('baselines', '实验设计：Baselines'),
    ('metrics', '实验设计：评价指标'),
    ('overallResults', '实验结果：总体结果'),
    ('coreEffect', '实验结果：核心贡献的效果'),
    ('ablation', '实验结果：方法消融'),
    ('inferenceSpeed', '实验结果：推理速度'),
    ('innovation1', '创新点 1'),
    ('innovation2', '创新点 2'),
    ('innovation3', '创新点 3'),
    ('innovation4', '创新点 4'),
    ('inspirationNote', '启发'),
    ('limitations', '局限性/待追问'),
    ('newIdeas', '可以推出的新研究思路'),
]

def convert_value(val):
    if val is None:
        return ''
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    if isinstance(val, (int, float)):
        # Excel dates may be stored as integers like 20260524
        if isinstance(val, int) and val > 20000101 and val < 21000101:
            s = str(val)
            return f"{s[:4]}-{s[4:6]}-{s[6:8]}"
        return val
    s = str(val).strip()
    # Also handle string dates like 20260524
    if s.isdigit() and len(s) == 8 and s.startswith('20'):
        return f"{s[:4]}-{s[4:6]}-{s[6:8]}"
    return s

def parse_sheet(ws, category_name):
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    header = rows[0]
    papers = []
    for row in rows[1:]:
        # Skip if all empty
        if not any(row):
            continue
        # Build mapping from header name to value
        row_dict = {}
        for h, v in zip(header, row):
            if h is not None:
                row_dict[str(h).strip()] = v

        uid = f"excel_{category_name}_{len(papers)}_{abs(hash(str(row))) % 0xFFFFFF:06x}"
        paper = {'id': uid, 'category': category_name}
        has_content = False
        for key, header_name in FIELD_MAP:
            val = row_dict.get(header_name)
            converted = convert_value(val)
            if converted not in ('', None):
                has_content = True
            # Handle numeric scores
            if key in ('relevance','novelty','evidence','inspiration','reproducibility'):
                try:
                    converted = float(converted) if converted != '' else 0
                except:
                    converted = 0
            paper[key] = converted
        if has_content and (paper.get('title') or paper.get('oneSentence') or paper.get('authors')):
            papers.append(paper)
    return papers

def main():
    xlsx_path = os.path.join(os.path.dirname(__file__), '..', '具身智能文献阅读记录表.xlsx')
    if not os.path.exists(xlsx_path):
        print(f"Error: {xlsx_path} not found")
        sys.exit(1)

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    all_papers = []

    for sheet_name, cat_name in CATEGORY_MAP.items():
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            papers = parse_sheet(ws, cat_name)
            print(f"Sheet '{sheet_name}': {len(papers)} papers")
            all_papers.extend(papers)
        else:
            print(f"Sheet '{sheet_name}' not found, skipping")

    # Also try 文献深读表-汇总
    if '文献深读表-汇总' in wb.sheetnames:
        ws = wb['文献深读表-汇总']
        papers = parse_sheet(ws, '接触丰富操作与灵巧控制')  # default if unknown
        # Only add if not already from category sheets (simple heuristic: same title)
        existing_titles = {p.get('title','') for p in all_papers}
        for p in papers:
            if p.get('title') not in existing_titles:
                all_papers.append(p)
        print(f"Sheet '文献深读表-汇总': added {len([p for p in papers if p.get('title') not in existing_titles])} new papers")

    output = {
        'version': 1,
        'exportedAt': datetime.now().isoformat(),
        'papers': all_papers
    }

    out_path = os.path.join(os.path.dirname(__file__), 'papers_from_excel.json')
    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"\nSaved {len(all_papers)} papers to {out_path}")

if __name__ == '__main__':
    main()
